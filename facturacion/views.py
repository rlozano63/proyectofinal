from django.db.models import Sum
from django.core import serializers
from django.shortcuts import render

from base.views import AjaxableResponseMixin
from base.models import distribuidor, cliente, producto

from django.http import JsonResponse

from django.views.decorators.csrf import csrf_exempt
import json

from facturacion.forms import *
from django.views.generic.edit import CreateView
from django.views.generic.list import ListView
from django.core.urlresolvers import reverse_lazy

from inventario.views import calcular_cantidad

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
# Create your views here.
import cStringIO as StringIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse

from django.views.generic import TemplateView

def fetch_resources(uri, rel):
    path = join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
    return path

class DemoPDFView(TemplateView):
    template_name = 'factura/imprimir.html'

    def render_to_response(self, context, **response_kwargs):
        pkfactura = self.kwargs['pkfactura']

        objfactura = factura.objects.get(pk=pkfactura)
        print (objfactura)
        objfactura.detalle = factura_detalle.objects.filter(factura=objfactura)
        print (len(objfactura.detalle))
        print (objfactura.detalle)

        context['objfactura'] = objfactura

        template = get_template(self.template_name)
        html = template.render(context)
        result = StringIO.StringIO()
        pdf = pisa.pisaDocument(
            StringIO.StringIO(html.encode("ISO-8859-1")),
            dest=result, link_callback=fetch_resources)
        if not pdf.err:
            return HttpResponse(result.getvalue(), content_type='application/pdf')
        return HttpResponse("Error: <pre>%s</pre>" % escape(html))

class FacturaList(LoginRequiredMixin,AjaxableResponseMixin,ListView):
	model = factura
	template_name = 'factura/listar.html'
	fields = "__all__"
	success_url = reverse_lazy('listar_factura')

	def get_context_data(self,**kwargs):
		context = super(FacturaList, self).get_context_data(**kwargs)
		return context

@login_required
def FacturaDetail(request, pkfactura):
	objfactura = factura.objects.get(pk=pkfactura)
	queryset = factura_detalle.objects.filter(factura=objfactura)
	context = {"factura":objfactura,"detalle":queryset}
	return render(request,"factura/detalle.html",context)


class FacturaCreation(LoginRequiredMixin,AjaxableResponseMixin,CreateView):
	model = factura
	template_name = 'factura/create.html'
	form_class = facturaForm
	#fields = "__all__"
	success_url = reverse_lazy('listar_productos')

	def get_context_data(self,**kwargs):
		context = super(FacturaCreation, self).get_context_data(**kwargs)
		context['form_detalle'] = facturaDetalleForm()
		return context

	def form_valid(self, form):
		deta = form.instance
		verificacion_ventas_cliente(deta.cliente.id)

		return super(FacturaCreation, self).form_valid(form)

class FacturaDetaleCreation(LoginRequiredMixin,AjaxableResponseMixin,CreateView):
	model = factura_detalle
	#template_name = 'factura/detalle_crear.html'
	#form_class = facturaForm
	fields = "__all__"
	success_url = reverse_lazy('listar_productos')
	def form_valid(self, form):
		tipo_movimiento = 2 # salida
		deta = form.instance
		#calcular_cantidad(deta.producto.pk,tipo_movimiento,deta.cantidad)

		deta.factura.valor_total += (deta.cantidad*deta.valor)
		deta.factura.save()

		return super(FacturaDetaleCreation, self).form_valid(form)




def verificacion_ventas_cliente(cliente_id = None):
	compra_minima_cliente = 0
	response = []
	if cliente_id:
		oclientes = [cliente.objects.get(id=cliente_id)]
	else:
		oclientes = cliente.objects.all()

	for ocliente in oclientes:
		total_ventas_cliente = factura.objects.filter(cliente=ocliente).aggregate(Sum("valor_total"))["valor_total__sum"]
		ocliente.ruta_activa = 1
		if total_ventas_cliente < compra_minima_cliente:
			ocliente.ruta_activa = 0
		ocliente.save()
	response = serializers.serialize('json', cliente.objects.all())
	return response

@csrf_exempt
@login_required
def url_verificacion_ventas_cliente(request):
	response = verificacion_ventas_cliente()
	return JsonResponse(response,safe=False)
from django.db.models import Sum
from django.db.models.functions import TruncMonth


@login_required

def reportes(request):
	clientes = cliente.objects.all()
	return render(request,"reportes/index.html",{"clientes":clientes})
@login_required

def ReporteVentasDistribuidor(request,id_distribuidor):
	pass

from django.db.models.functions import TruncMonth
from django.db.models import Sum, Count

from openpyxl import Workbook
from django.contrib.humanize.templatetags.humanize import intcomma

def ReporteMasVendidosExcel(context):
	wb = Workbook()
	ws = wb.active

	title = 'Productos Vendidos'
	print(context['config']['cliente'])
	if context['config']['cliente'] != None:
		title = title + " a {0}".format(context['config']['cliente'].nombre)
	title = title + " desde {0} hasta {1}".format(context['config']['fini'], context['config']['ffin'])

	ws['A1'] = title
	ws.merge_cells('A1:E1')
	ws['A3'] = 'id'
	ws['B3'] = 'Producto'
	ws['C3'] = 'Cantidad'
	cont=4
	for producto in context['productos']:
		ws.cell(row=cont,column=1).value = producto.pk
		ws.cell(row=cont,column=2).value = producto.nombre
		ws.cell(row=cont,column=3).value = producto.cantidad
		cont = cont + 1

	nombre_archivo ="ReporteMasVendidos.xlsx"
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response

def ReporteVentasTotalExcel(context):
	wb = Workbook()
	ws = wb.active

	title = 'Reporte de Ventas Totales'
	print(context['config']['cliente'])
	if context['config']['cliente'] != None:
		title = title + " a {0}".format(context['config']['cliente'].nombre)
	title = title + " desde {0} hasta {1}".format(context['config']['fini'], context['config']['ffin'])

	ws['A1'] = title
	ws.merge_cells('A1:E1')
	ws['A3'] = 'id'
	ws['B3'] = 'Fecha'
	ws['C3'] = 'Total'
	cont=4

	for factura in context['facturas']:
		ws.cell(row=cont,column=1).value = factura.pk
		ws.cell(row=cont,column=2).value = factura.fecha_creacion.strftime('%Y-%m-%d')
		ws.cell(row=cont,column=3).value = "$" + intcomma(factura.valor_total)
		cont = cont + 1

	ws.cell(row=cont,column=1).value = 'Total'
	ws.cell(row=cont,column=3).value = "$" + intcomma(context['vttotal'])

	nombre_archivo ="ReporteVentasTotales.xlsx"
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response

@login_required
def ReporteVentasTotal(request):


	fini = request.GET.get("fini")
	ffin = request.GET.get("ffin")

	formato = request.GET.get("formato", "html")

	cliente_id = request.GET.get("cliente")
	clienteobj = None
	if ( cliente_id == None or cliente_id == '' ):
		facturas = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin)
		vttotal = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin).aggregate(Sum("valor_total"))["valor_total__sum"]

		facturasMes = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin).annotate(month=TruncMonth('fecha_creacion')).values('month').annotate(c=Count('id')).annotate(sum=Sum('valor_total')).order_by()
	else:
		clienteobj = cliente.objects.filter(id=cliente_id).get()
		facturas = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin,cliente__id=cliente_id)
		vttotal = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin,cliente__id=cliente_id).aggregate(Sum("valor_total"))["valor_total__sum"]
		facturasMes = factura.objects.filter(fecha_creacion__date__gte=fini,fecha_creacion__date__lte=ffin,cliente__id=cliente_id).annotate(month=TruncMonth('fecha_creacion')).values('month').annotate(c=Count('id')).annotate(sum=Sum('valor_total')).order_by()


	for fac in facturas:
		fac.detalle = factura_detalle.objects.filter(factura = fac)

	data = []
	for facMes in facturasMes:
		data.append([facMes['month'].strftime("%d/%m/%y"), facMes['sum']])

	context = {
		"facturas":facturas,
		"vttotal":vttotal,
		"json":json.dumps(data),
		"config":{
			"cliente": clienteobj,
			"fini": fini,
			"ffin": ffin,
		}
	}

	if formato == 'excel':
		return ReporteVentasTotalExcel(context)
	else:
		return render(request,"reportes/ventasTotal.html",context)

def ReporteMasVendidos(request):
	fini = request.GET.get("fini")
	ffin = request.GET.get("ffin")

	formato = request.GET.get("formato", "html")

	cliente_id = request.GET.get("cliente")
	clienteobj = None
	productos = producto.objects.all()

	data = []
	for pro in productos:
		if ( cliente_id == None or cliente_id == '' ):

			pro.detalles = factura_detalle.objects.filter(producto = pro)
		else:
			clienteobj = cliente.objects.filter(id=cliente_id).get
			pro.detalles = factura_detalle.objects.filter(producto = pro,factura__cliente__id=cliente_id)

		pro.cantidad = len(pro.detalles)
		data.append([pro.nombre,pro.cantidad])

	context = {
		"productos":productos,
		"json":json.dumps(data),
		"config":{
			"cliente": clienteobj,
			"fini": fini,
			"ffin": ffin,
		}
	}

	if formato == 'excel':
		return ReporteMasVendidosExcel(context)
	else:
		return render(request,"reportes/masVendidos.html",context)




